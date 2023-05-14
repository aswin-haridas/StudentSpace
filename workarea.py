import openpyxl
import random

# Create a workbook object
workbook = openpyxl.Workbook()

# Create a sheet object
sheet = workbook.active

# Write column headers
sheet['A1'] = 'Student ID'
sheet['B1'] = 'Date'
sheet['C1'] = 'Attendance Status'

# Create a list of student IDs
student_ids = ['CS{:03d}'.format(i) for i in range(1, 51)]

# Create a dictionary of dates and attendance status
attendance_data = {}

# Generate attendance data for each day in January
for day in range(1, 32):
    date = f'{day:02}/01/2022'
    for student_id in student_ids:
        # Generate random attendance status with priority to 1
        if random.randint(1, 10) <= 8:
            attendance_data[(student_id, date)] = 1
        else:
            attendance_data[(student_id, date)] = 0

# Write attendance data to Excel sheet
row = 2  # start writing from row 2
for (student_id, date), attendance_status in attendance_data.items():
    sheet.cell(row=row, column=1).value = student_id
    sheet.cell(row=row, column=2).value = date
    sheet.cell(row=row, column=3).value = attendance_status
    row += 1

# Save the workbook
workbook.save('attendance.xlsx')
