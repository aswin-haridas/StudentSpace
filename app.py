from flask import Flask, render_template, request, redirect, url_for, session, jsonify
from openpyxl import load_workbook
import sqlite3
import os

app = Flask(__name__)
app.secret_key = "your_secret_key_here"


@app.context_processor
def common_icons():
    admin = "/static/assets/admin.png"
    student = "/static/assets/student.png"
    faculty = "/static/assets/faculty.png"
    search = "/static/assets/search.png"
    notification = "/static/assets/notification.png"
    settings = "/static/assets/settings.png"
    editimage="/static/assets/edit.png"

    return dict(
        search=search,
        notification=notification,
        settings=settings,
        admin=admin,
        faculty=faculty,
        student=student,
        editimage=editimage
    )


@app.route("/", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        user_type_form = request.form.get("user-type")
        username = request.form.get("username")
        password = request.form.get("password")

        conn = sqlite3.connect("database.db")
        cursor = conn.cursor()
        cursor.execute(
            "SELECT id, role, username FROM users WHERE username=? AND password=?",
            (username, password),
        )
        result = cursor.fetchone()

        if result is not None:
            user_id, user_type, name = result
            session["user_id"] = user_id
            session["user_type"] = user_type
            session["username"] = username
            session["name"] = name

            if user_type_form == user_type:
                return redirect(url_for("home"))
        error_message = "Invalid username or password"
        return render_template("error.html", error=error_message)

    return render_template("login.html")


@app.route("/home", endpoint="home")
def home():
    user_id = session.get("user_id")
    user_type = session.get("user_type")
    username = session.get("username")
    name = session.get("name")
    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()
    if user_type == "student":
        cursor.execute("SELECT name FROM studentlist WHERE id=?", (user_id,))
    elif user_type == "faculty":
        cursor.execute("SELECT name FROM facultylist WHERE id=?", (user_id,))
    elif user_type == "admin":
        cursor.execute("SELECT name FROM facultylist WHERE id=?", (user_id,))

    result = cursor.fetchone()
    name = None
    if result is not None:
        name = result[0]

    session["name"] = name
    cursor.execute("SELECT status FROM attendance WHERE id=?", (user_id,))
    attendance_data = [status[0] for status in cursor.fetchall()]

    total_attendance = len(attendance_data)
    present_attendance = sum(1 for status in attendance_data if status == "Present")

    attendance_percentage = 0
    if total_attendance != 0:
        attendance_percentage = (present_attendance / total_attendance) * 100

    conn.close()

    return render_template(
        "home.html",
        username=username,
        name=name,
        user_type=user_type,
        attendance_percentage=attendance_percentage,
    )

def connect_to_database():
    conn = sqlite3.connect("database.db")
    return conn

@app.route("/profile")
def profile():
    user_id = session.get("user_id")
    user_type = session.get("user_type")

    if user_id is None:
        return "User ID not found"

    conn = connect_to_database()
    cursor = conn.cursor()

    if user_type == "student":
        cursor.execute("SELECT id, name, dob, email, course, contact, address, pfp FROM studentlist WHERE id=?", (user_id,))
        user_info = cursor.fetchone()
        if user_info is None:
            conn.close()
            return "User not found"

        id, name, dob, email, course, contact, address, pfp = user_info
        conn.close()
        return render_template(
            "profile.html",
            id=id,
            user_type=user_type,
            name=name,
            dob=dob,
            email=email,
            pfp=pfp,
            course=course,
            contact=contact,
            address=address,
        )
    elif user_type == "faculty":
        cursor.execute("SELECT id, name, dob, email, department, contact, address, pfp FROM facultylist WHERE id=?", (user_id,))
        user_info = cursor.fetchone()
        if user_info is None:
            conn.close()
            return "User not found"

        id, name, dob, email, department, contact, address, pfp = user_info
        conn.close()
        return render_template(
            "profile.html",
            id=id,
            user_type=user_type,
            name=name,
            dob=dob,
            email=email,
            pfp=pfp,
            course=department,
            contact=contact,
            address=address,
        )

    return "User not found"


@app.route("/studentlist")
def get_student_list():
    conn = sqlite3.connect("database.db")
    c = conn.cursor()
    c.execute("SELECT id, name ,email FROM studentlist")
    student_list = c.fetchall()
    conn.close()
    return jsonify(student_list)

@app.route("/facultylist")
def get_faculty_list():
    conn = sqlite3.connect("database.db")
    c = conn.cursor()
    c.execute("SELECT id, name, department FROM facultylist")  # Adjust the table name as needed
    faculty_list = c.fetchall()
    conn.close()
    return jsonify(faculty_list)



@app.route("/attendancepercentage")
def attendancepercentage():
    user_id = session.get("user_id")
    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()

    cursor.execute("SELECT SUM(status) FROM attendance WHERE id=?", (user_id,))
    sum_status = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM attendance WHERE id=?", (user_id,))
    num_rows = cursor.fetchone()[0]

    attendance_percentage = (sum_status / num_rows) * 100 if num_rows > 0 else 0
    attendance_percentage = round(attendance_percentage, 1)

    cursor.close()
    conn.close()

    return jsonify(attendance_percentage)


@app.route("/grades")
def get_student_grades():
    user_id = session.get("user_id")
    conn = sqlite3.connect("database.db")
    c = conn.cursor()
    c.execute("SELECT s1, s2, s3, s4, s5, s6 FROM grades WHERE id = ?", (user_id,))
    data = c.fetchone()
    conn.close()
    if data:
        grade_data = {
            "s1": data[0],
            "s2": data[1],
            "s3": data[2],
            "s4": data[3],
            "s5": data[4],
            "s6": data[5],
        }
        return jsonify(grade_data)
    else:
        return jsonify({})

@app.route('/tasks', methods=['GET'])
def tasks():
    conn = sqlite3.connect('database.db')
    cursor = conn.cursor()
    
    cursor.execute('SELECT * FROM tasks')
    tasks = cursor.fetchall()
    
    conn.close()
    
    task_list = []
    for task in tasks:
        task_dict = {
            'id': task[0],
            'task_name': task[1],
            'due_date': task[2],
            'status': task[3]
        }
        task_list.append(task_dict)
    
    return jsonify({'tasks': task_list})

@app.route('/add_task', methods=['POST'])
def add_task():
    task_name = request.form.get('task_name')
    due_date = request.form.get('due_date')
    status = request.form.get('status')

    # Connect to the SQLite database
    connection = sqlite3.connect('database.db')
    cursor = connection.cursor()

    # Insert the new task into the tasks table
    cursor.execute('INSERT INTO tasks (task_name, due_date, status) VALUES (?, ?, ?)', (task_name, due_date, status))
    connection.commit()

    # Fetch the newly added task
    cursor.execute('SELECT * FROM tasks WHERE task_name = ?', (task_name,))
    new_task = cursor.fetchone()

    # Close the database connection
    connection.close()

    # Return the new task as JSON response
    return jsonify(new_task)

@app.route("/perfomance")
def perfomance():
    user_id = session.get("user_id")
    user_type = session.get("user_type")
    name = session.get("name")
    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM grade WHERE id = ?", (user_id,))
    marks = cursor.fetchone()
    cursor.execute("SELECT * FROM grades")
    grades = cursor.fetchall()
    conn.close()
    return render_template(
        "perfomance.html", user_type=user_type, name=name, marks=marks, grades=grades
    )


@app.route("/upload")
def upload():
    return render_template("upload-faculty.html")


@app.route("/process", methods=["POST"])
def process():
    file = request.files["file"]
    wb = load_workbook(file)
    sheet = wb.active
    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS attendance (
            date,
            student_id,
            attendance_status
        )
    """
    )
    for row in sheet.iter_rows(min_row=2):
        date, student_id, attendance_status = row
        cursor.execute(
            "INSERT INTO attendance VALUES (?, ?, ?)",
            (date, student_id, attendance_status),
        )
    conn.commit()
    conn.close()
    message = "File uploaded successfully!"
    return render_template("upload.html", message=message)


@app.route("/attendance", methods=["GET", "POST"])
def attendance():
    user_id = session.get("user_id")
    user_type = session.get("user_type")
    name = session.get("name")
    if request.method == "POST":
        selected_month = request.form["month"]
    else:
        selected_month = "January"  # Default month

    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()

    cursor.execute(
        "SELECT status FROM attendance WHERE id=? AND month=?",
        (user_id, selected_month),
    )
    attendance_records = cursor.fetchall()

    conn.close()

    if len(attendance_records) == 0:
        return f"No attendance records found for user ID {user_id} in {selected_month}."

    return render_template(
        "attendance.html",
        user_type=user_type,
        name=name,
        attendance_records=attendance_records,
        selected_month=selected_month,
    )


@app.route("/classAttendance", methods=["GET", "POST"])
def classAttendance():
    user_type = session.get("user_type")
    name = session.get("name")
    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()

    selected_day = request.args.get("day")
    if selected_day is None:
        selected_day = "1"

    cursor.execute("SELECT DISTINCT date FROM attendance")
    dates = cursor.fetchall()
    days = [str(date[0]) for date in dates]

    cursor.execute(
        "SELECT id, name, status FROM attendance WHERE date=?",
        (selected_day,),
    )
    rows = cursor.fetchall()

    attendance_data = {}
    for row in rows:
        student_id = row[0]
        attendance_name = row[1]
        attendance_status = row[2]
        attendance_data[student_id] = {
        'name': attendance_name,
        'status': attendance_status
    }
    conn.close()
    return render_template(
        "attendance.html",
        user_type=user_type,
        attendance_data=attendance_data,
        name=name,
        days=days,
        selected_day=selected_day,
    )

@app.route("/courses")
def courses():
    user_type = session.get("user_type")
    name = session.get("name")
    
    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()
    cursor.execute("SELECT course_id, course_name, instructor, credits, course_image FROM courses")
    courses_data = cursor.fetchall()
    conn.close()

    courses = []
    for course in courses_data:
        course_dict = {
            "course_id": course[0],
            "course_name": course[1],
            "instructor": course[2],
            "credits": course[3],
            "course_image": course[4]
        }
        courses.append(course_dict)
    
    return render_template(
        "courses.html",
        courses=courses,
        user_type=user_type,
        name=name,
    )

@app.route("/coursemgmt")
def coursemgmt():
    user_type = session.get("user_type")
    name = session.get("name")
    
    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()
    cursor.execute("SELECT name, course FROM faculties")
    coursemgmt = [
        dict(name=row[0], course=row[1])
        for row in cursor.fetchall()
    ]
    conn.close()
    
    
    return render_template(
        "coursemgmt.html",
        coursemgmt=coursemgmt,
        user_type=user_type,
        name=name,
    )
 
@app.route('/role_management')
def role_management():
    name = session.get("name")
    connection = connect_to_database()
    cursor = connection.cursor()
    cursor.execute('SELECT * FROM roles')
    roles_data = cursor.fetchall()
    connection.close()
    return render_template('rolemgmt.html', roles_data=roles_data, name=name)


@app.route('/user_management')
def user_management():
    name=session.get("name")
    connection = connect_to_database()
    cursor = connection.cursor()
    cursor.execute('SELECT * FROM users')
    users_data = cursor.fetchall()
    connection.close()
    return render_template('usermgmt.html', users_data=users_data,name=name)



@app.route("/notes")
def notes():
    user_type = session.get("user_type")
    name = session.get("name")
    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()
    cursor.execute("SELECT title, content, uploaded_by, file_url FROM notes")
    notes = [
        dict(title=row[0], content=row[1], uploaded_by=row[2], file_url=row[3])
        for row in cursor.fetchall()
    ]

    conn.close()
    return render_template(
        "notes.html",
        notes=notes,
        user_type=user_type,
        name=name,
    )


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


if __name__ == "__main__":
    app.run(debug=True)
